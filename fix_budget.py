from openpyxl import Workbook
import random

wb = Workbook()

# grab the active worksheet
ws = wb.active

#количество траифка по позициям
traffic = [1, 0.85, 0.75, 0.65, 0.06, 0.05, 0.04, 0.03, 0.02, 0, 0]

#целевая стоимость конверсии
click_value = 800

#недельный бюджет
week_budget = 30000


ver_conv = [0]*50
amount_traff = [0]*50

    
cost_click = [25]*50 #массив ставок по ключам. начальная ставка 20 руб.
base_cost_click = 20 #контрольная фиксированная ставка для всех ключей - 20 руб
cost_pos = [[0]*9 for i in range(50)] #массив стоимости позиций
all_count_conv = 0
all_money = 0
all_money_fix = 0
u = 0
l = 50
count_conv = [0]*50
all_key_conv = [0]*50
all_key_costs = [0]*50
costs_per_key = [0]*50
full_money = 0
full_conv = 0

full_money_fix = 0
full_conv_fix = 0
full_costs_per_key = [0]*50
full_conv_per_key = [0]*50

full_full_money = 0
full_full_conv = 0
        
full_full_money_fix = 0
full_full_conv_fix = 0

cost_click = [25]*50
random.seed
for q in range(10):
    
    
    #задаем вероятности конверсий и количетсво трафика ключам
    for i in range(50): 
        ver_conv[i] = random.randint(1, 50) #процент конверсии умноженный на 10
        amount_traff[i] = random.randint(1, 50)

    for x in range (1, 5):
        all_count_conv = 0
        all_count_conv_fix = 0
        money = 0
        money_fix = 0
        traff_period = [0]*50

        #генерируем стоимость позиций для ключей
        for v in range(1, 50):
            cost_pos[v][8] = round(random.uniform(1, 3), 2)
            cost_pos[v][7] = cost_pos[v][8] + round(random.uniform(0.5, 2), 2)
            cost_pos[v][6] = cost_pos[v][7] + round(random.uniform(0.5, 2), 2)
            cost_pos[v][5] = cost_pos[v][6] + round(random.uniform(0.5, 2), 2)
            cost_pos[v][4] = cost_pos[v][5] + round(random.uniform(0.5, 3), 2)
            cost_pos[v][3] = cost_pos[v][4] + round(random.uniform(4, 10), 2)
            cost_pos[v][2] = cost_pos[v][3] + round(random.uniform(2, 10), 2)
            cost_pos[v][1] = cost_pos[v][2] + round(random.uniform(1, 15), 2)
            cost_pos[v][0] = cost_pos[v][1] + round(random.uniform(2, 10), 2)
            
            #записываем в excel цены 
            for i in range(1, 10):
                ws.cell(row=v+u, column=i).value = cost_pos[v][i-1]

        num_pos= [10]*50
        cpc = [0]*50
        for v in range(1, 50):           
            #ищем максимальную доступную позицию и стоимость клика для оптимизатора 
            for i in range(9):
                if cost_click[v] == 0:
                    num_pos[v] = 10
                    cpc[v] = 0
                    break
                if cost_pos[v][i] <= cost_click[v]:
                    cpc[v] = cost_pos[v][i]+0.01
                    num_pos[v] = i
                    break

        #до 10 итераций аукциона за день при ограниченном бюджете
        q = 1
        day_budget = 0
        max_day_budget = 5000
        period_amount_traff = [[0]*10 for i in range(50)]
        day_traffic = [0]*50
        day_costs_per_key = [0]*50

        #максимальное количетсво трафика у ключа за период
        for i in range(50):
            if amount_traff[i] >= 10:
                for j in range(amount_traff[i] // 10):
                    period_amount_traff[i][j] = amount_traff[i] // 10
                for j in range(amount_traff[i] % 10):
                    period_amount_traff[i][j] += amount_traff[i] % 10
            else:
                for j in range(amount_traff[i] % 10):
                    period_amount_traff[i][j] += amount_traff[i] % 10
                    

        j = 0
        while (day_budget < max_day_budget * 0.7) and (q <= 10):
            for i in range(50):
                #считаем количество трафика за период для оптимизатора
                traff_period[i] = round(period_amount_traff[i][j] * traffic [num_pos[i]], 0)
                day_traffic[i] += traff_period[i]

                #расходы по ключу оптимизатора
                costs_per_key[i] = traff_period[i] * cpc[i]
                full_costs_per_key[i] += costs_per_key[i]
                day_costs_per_key[i] += costs_per_key[i]

                #дневной расход
                day_budget += costs_per_key[i]

            q += 1
            j += 1


        count_conv = [0] * 50
        for v in range(1, 50):
            #количество конверсий оптимизатора
            for j in range(int(day_traffic[v])):
                if random.randint(1, 1000) <= ver_conv[v]:
                    count_conv[v] += 1
        
            all_count_conv += count_conv[v]
            full_conv_per_key[v] += count_conv[v]

            conv_cost = [0]*50
            #стоимость конверсии по ключам с опитимизатором
            if count_conv[v] > 0:
                conv_cost[v] = day_costs_per_key[v] / count_conv[v]
                    
            #расходы за период
            money += day_costs_per_key[v]
                
            #оптимизированная ставка
            ws.cell(row=v+u, column=18).value = cpc[v]
            ws.cell(row=v+u, column=19).value = num_pos[v]
            ws.cell(row=v+u, column=20).value = traff_period[v]
            ws.cell(row=v+u, column=21).value = day_costs_per_key[v]
            ws.cell(row=v+u, column=22).value = count_conv[v]
            ws.cell(row=v+u, column=23).value = conv_cost[v]
            ws.cell(row=v+u, column=24).value = cost_click[v]
            ws.cell(row=v+u, column=28).value = full_costs_per_key[v]
            ws.cell(row=v+u, column=29).value = full_conv_per_key[v]
            if full_conv_per_key[v] > 0:
                ws.cell(row=v+u, column=30).value = full_costs_per_key[v] / full_conv_per_key[v]


        #считаем расходы и конверсии у оптимизатора
        if all_count_conv > 0:
            cost_conv_period = money / all_count_conv
        else:
            cost_conv_period = money
        all_money += money

        ws.cell(row=l, column=21).value = money
        ws.cell(row=l, column=22).value = all_count_conv
        ws.cell(row=l, column=23).value = cost_conv_period
        ws.cell(row=l, column=26).value = q
        

        full_money += money
        full_conv += all_count_conv
        
        full_money_fix += money_fix
        full_conv_fix += all_count_conv_fix

        u += 52
        l += 52

        full_full_money += full_money
        full_full_conv += full_conv
        
        full_full_money_fix += full_money_fix
        full_full_conv_fix += full_conv_fix
        
    #выставляем ставки
    for i in range(1, 50):
        if full_conv_per_key[i] > 0:
            if full_costs_per_key[i] / full_conv_per_key[i] < click_value * 0.9:
                cost_click[i] = cost_click[i] + cost_click[i]*0.1
            else:
                cost_click[i] = cost_click[i] - cost_click[i]*0.1
                    
            if full_costs_per_key[i] / full_conv_per_key[i] > click_value*1.5:
                cost_click[i] = 0
        else:
            if full_costs_per_key[i] > click_value*2:
                cost_click[i] = 0

          

    

ws.cell(row=1, column=32).value = full_money
ws.cell(row=1, column=33).value = full_conv
ws.cell(row=1, column=34).value = full_money / full_conv

#ws.cell(row=2, column=32).value = full_full_money_fix
#ws.cell(row=2, column=33).value = full_full_conv_fix
#ws.cell(row=2, column=34).value = full_full_money_fix / full_full_conv_fix

# Save the file
wb.save("sample2.xlsx")

print ("Done")
